from __future__ import annotations

import json

import pandas as pd
from openpyxl import load_workbook

from nixtla_scaffold import ForecastSpec, run_forecast
from nixtla_scaffold.cli import main
from nixtla_scaffold.comparisons import (
    FORECAST_COMPARISON_SCHEMA_VERSION,
    compare_forecasts,
    write_forecast_comparison,
)


def test_compare_forecasts_aligns_external_forecast_without_scoring(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "ds": forecast["ds"],
            "yhat": forecast["yhat"] + 10,
            "model": ["Finance plan"] * len(forecast),
            "source_id": ["fpna_workbook"] * len(forecast),
            "scenario_name": ["base"] * len(forecast),
        }
    )

    result = compare_forecasts(run_dir, external)

    comparison = result.comparison
    assert set(comparison["comparison_status"]) == {"aligned"}
    assert comparison["comparison_delta_yhat"].tolist() == [10.0, 10.0]
    assert comparison["is_comparison_scoreable"].tolist() == [False, False]
    assert comparison["external_is_actual"].tolist() == [False, False]
    assert comparison["comparison_warning"].str.contains("Directional triangulation only").all()
    assert comparison["comparison_evidence_status"].tolist() == ["future_only_unscored", "future_only_unscored"]
    assert not any("accuracy" in column.lower() or column.lower() in {"mae", "rmse", "mape"} for column in comparison.columns)

    assert result.summary.loc[0, "aligned_rows"] == 2
    assert result.summary.loc[0, "rows_with_both_forecasts"] == 2
    assert result.summary.loc[0, "avg_abs_delta_yhat_vs_scaffold"] == 10.0
    assert not bool(result.summary.loc[0, "is_comparison_scoreable"])
    assert result.manifest["schema_version"] == FORECAST_COMPARISON_SCHEMA_VERSION
    assert result.manifest["alignment"]["status"] == "aligned"
    assert result.manifest["alignment"]["rows_aligned"] == 2
    assert result.manifest["alignment"]["rows_comparable"] == 2
    assert result.manifest["external"]["evidence_status_distribution"] == {"future_only_unscored": 2}
    assert "Deltas are not residuals" in "\n".join(result.manifest["guardrails"])


def test_compare_forecasts_preserves_same_model_scenario_multi_source_lineage(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.concat(
        [
            pd.DataFrame(
                {
                    "unique_id": forecast["unique_id"],
                    "ds": forecast["ds"],
                    "yhat": forecast["yhat"] + offset,
                    "model": ["Plan"] * len(forecast),
                    "source_id": [source_id] * len(forecast),
                    "scenario_name": ["base"] * len(forecast),
                }
            )
            for source_id, offset in [("fpna_workbook", 5), ("sales_plan", 10)]
        ],
        ignore_index=True,
    )

    result = compare_forecasts(run_dir, external)

    assert set(result.comparison["comparison_status"]) == {"aligned"}
    assert set(result.comparison["external_source_id"]) == {"fpna_workbook", "sales_plan"}
    assert len(result.comparison) == len(forecast) * 2
    summary = result.summary.sort_values("external_source_id").reset_index(drop=True)
    assert summary["external_source_id"].tolist() == ["fpna_workbook", "sales_plan"]
    assert summary["aligned_rows"].tolist() == [len(forecast), len(forecast)]
    assert result.manifest["alignment"]["rows_external"] == len(external)
    assert result.manifest["summary_rows"] == 2


def test_write_forecast_comparison_outputs_artifacts(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    original_forecast = (run_dir / "forecast.csv").read_bytes()
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "ds": forecast["ds"],
            "yhat": forecast["yhat"],
            "model": ["Plan"] * len(forecast),
            "source_id": ["finance_export"] * len(forecast),
        }
    )

    result = write_forecast_comparison(run_dir, external, output_dir=tmp_path / "comparison")

    assert result.manifest["output_dir"] == str(tmp_path / "comparison")
    assert (tmp_path / "comparison" / "forecast_comparison.csv").exists()
    assert (tmp_path / "comparison" / "forecast_comparison_summary.csv").exists()
    assert (tmp_path / "comparison" / "forecast_comparison.xlsx").exists()
    assert (tmp_path / "comparison" / "comparison_report.html").exists()
    assert (tmp_path / "comparison" / "comparison_llm_context.json").exists()
    assert (run_dir / "forecast.csv").read_bytes() == original_forecast
    manifest = json.loads((tmp_path / "comparison" / "comparison_manifest.json").read_text(encoding="utf-8"))
    assert manifest["outputs"]["comparison"] == "forecast_comparison.csv"
    assert manifest["outputs"]["workbook"] == "forecast_comparison.xlsx"
    assert manifest["outputs"]["html_report"] == "comparison_report.html"
    assert manifest["outputs"]["llm_context"] == "comparison_llm_context.json"
    assert manifest["comparison_scope"] == "directional_unscored"


def test_write_forecast_comparison_workbook_html_and_llm_context_are_readable(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    original_files = {
        name: (run_dir / name).read_bytes()
        for name in ["forecast.csv", "report.html", "llm_context.json"]
        if (run_dir / name).exists()
    }
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "ds": forecast["ds"],
            "yhat": forecast["yhat"] + [3, -2],
            "model": ["Plan"] * len(forecast),
            "source_id": ["finance_export"] * len(forecast),
            "scenario_name": ["base"] * len(forecast),
        }
    )

    result = write_forecast_comparison(run_dir, external, output_dir=tmp_path / "comparison")

    workbook = load_workbook(tmp_path / "comparison" / "forecast_comparison.xlsx", read_only=True)
    assert {"Summary", "Comparison Rows", "Alignment", "Guardrails", "External Forecasts"}.issubset(set(workbook.sheetnames))
    assert workbook["Summary"]["A1"].value == "external_model"
    assert workbook["Comparison Rows"]["A1"].value == "unique_id"
    assert workbook["Comparison Rows"].max_row == len(result.comparison) + 1

    html = (tmp_path / "comparison" / "comparison_report.html").read_text(encoding="utf-8")
    assert "Directional triangulation only" in html
    assert "Summary by external model/source/scenario" in html
    assert "<table" in html

    llm_context = json.loads((tmp_path / "comparison" / "comparison_llm_context.json").read_text(encoding="utf-8"))
    assert llm_context["schema_version"] == "nixtla_scaffold.forecast_comparison.llm_context.v1"
    assert llm_context["comparison_scope"] == "directional_unscored"
    assert llm_context["is_comparison_scoreable"] is False
    assert llm_context["external_is_actual"] is False
    assert "Deltas are not residuals" in "\n".join(llm_context["guardrails"])
    assert llm_context["artifact_index"]["workbook"] == "forecast_comparison.xlsx"

    for name, original in original_files.items():
        assert (run_dir / name).read_bytes() == original


def test_compare_forecasts_no_overlap_writes_gap_diagnostics(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    external = pd.DataFrame(
        {
            "unique_id": ["Revenue", "Revenue"],
            "ds": ["2026-12-31", "2027-01-31"],
            "yhat": [200, 210],
            "model": ["Plan", "Plan"],
            "source_id": ["finance_export", "finance_export"],
        }
    )

    result = compare_forecasts(run_dir, external)

    assert result.manifest["alignment"]["status"] == "no_overlap"
    assert result.manifest["alignment"]["rows_aligned"] == 0
    assert {"scaffold_only", "external_only"}.issubset(set(result.comparison["comparison_status"]))
    assert result.comparison["comparison_delta_yhat"].isna().all()


def test_compare_forecasts_flags_cutoff_mismatch_without_scoring(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "cutoff": ["2025-07-31"] * len(forecast),
            "ds": forecast["ds"],
            "yhat": forecast["yhat"],
            "model": ["Historical plan"] * len(forecast),
            "source_id": ["finance_snapshot"] * len(forecast),
        }
    )

    result = compare_forecasts(run_dir, external)

    assert set(result.comparison["comparison_status"]) == {"mismatch_cutoff"}
    assert set(result.comparison["comparison_evidence_status"]) == {"historical_cutoff_labeled_unscored"}
    assert result.comparison["is_comparison_scoreable"].eq(False).all()
    assert result.comparison["comparison_warning"].str.contains("Forecast origins differ").all()
    assert result.comparison["comparison_delta_yhat"].isna().all()
    assert result.summary.loc[0, "aligned_rows"] == 0
    assert result.summary.loc[0, "rows_with_both_forecasts"] == len(forecast)
    assert result.summary.loc[0, "mismatch_cutoff_rows"] == len(forecast)
    assert pd.isna(result.summary.loc[0, "avg_delta_yhat_vs_scaffold"])
    assert pd.isna(result.summary.loc[0, "avg_abs_delta_yhat_vs_scaffold"])
    assert result.manifest["alignment"]["status"] == "cutoff_mismatch_only"
    assert result.manifest["alignment"]["rows_aligned"] == 0
    assert result.manifest["alignment"]["rows_comparable"] == 0
    assert result.manifest["alignment"]["rows_with_both_forecasts"] == len(forecast)
    assert result.manifest["alignment"]["mismatch_cutoff_rows"] == len(forecast)

    out_dir = result.to_directory(tmp_path / "cutoff_comparison")
    workbook = load_workbook(out_dir / "forecast_comparison.xlsx", read_only=True)
    headers = [cell.value for cell in workbook["Comparison Rows"][1]]
    status_idx = headers.index("comparison_status") + 1
    delta_idx = headers.index("comparison_delta_yhat") + 1
    warning_idx = headers.index("comparison_warning") + 1
    assert workbook["Comparison Rows"].cell(row=2, column=status_idx).value == "mismatch_cutoff"
    assert workbook["Comparison Rows"].cell(row=2, column=delta_idx).value is None
    assert "Forecast origins differ" in workbook["Comparison Rows"].cell(row=2, column=warning_idx).value


def test_compare_forecasts_counts_distinct_external_cutoffs_in_manifest(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.concat(
        [
            pd.DataFrame(
                {
                    "unique_id": forecast["unique_id"],
                    "cutoff": [cutoff] * len(forecast),
                    "ds": forecast["ds"],
                    "yhat": forecast["yhat"] + offset,
                    "model": ["Historical plan"] * len(forecast),
                    "source_id": ["finance_snapshot"] * len(forecast),
                    "scenario_name": ["base"] * len(forecast),
                }
            )
            for cutoff, offset in [("2025-07-31", 2), ("2025-08-31", 4)]
        ],
        ignore_index=True,
    )

    result = compare_forecasts(run_dir, external)

    assert len(result.external_forecasts) == len(external)
    assert len(result.comparison) == len(external)
    assert set(result.comparison["comparison_status"]) == {"aligned", "mismatch_cutoff"}
    assert result.manifest["alignment"]["rows_external"] == len(external)
    assert result.manifest["alignment"]["rows_aligned"] == len(forecast)
    assert result.manifest["alignment"]["mismatch_cutoff_rows"] == len(forecast)
    assert result.manifest["alignment"]["status"] == "partially_aligned_with_cutoff_warnings"


def test_compare_forecasts_flags_unknown_scaffold_origin_for_cutoff_rows(tmp_path) -> None:
    run_dir = _write_small_run(tmp_path)
    (run_dir / "manifest.json").unlink()
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external = pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "cutoff": ["2025-07-31"] * len(forecast),
            "ds": forecast["ds"],
            "yhat": forecast["yhat"] + 3,
            "model": ["Historical plan"] * len(forecast),
            "source_id": ["finance_snapshot"] * len(forecast),
        }
    )

    result = compare_forecasts(run_dir, external)

    assert set(result.comparison["comparison_status"]) == {"unknown_cutoff_origin"}
    assert result.comparison["comparison_delta_yhat"].isna().all()
    assert result.comparison["comparison_warning"].str.contains("scaffold forecast origin could not be verified").all()
    assert result.summary.loc[0, "aligned_rows"] == 0
    assert result.summary.loc[0, "unknown_cutoff_origin_rows"] == len(forecast)
    assert result.manifest["inputs"]["scaffold_forecast_origin_status"] == "unknown_required_for_external_cutoffs"
    assert result.manifest["alignment"]["status"] == "unknown_cutoff_origin_only"
    assert result.manifest["alignment"]["rows_aligned"] == 0
    assert result.manifest["alignment"]["unknown_cutoff_origin_rows"] == len(forecast)


def test_compare_cli_writes_artifacts(tmp_path, capsys) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast = pd.read_csv(run_dir / "forecast.csv")
    external_path = tmp_path / "finance_plan.csv"
    pd.DataFrame(
        {
            "unique_id": forecast["unique_id"],
            "ds": forecast["ds"],
            "yhat": forecast["yhat"] + 1,
        }
    ).to_csv(external_path, index=False)

    exit_code = main(
        [
            "compare",
            "--run",
            str(run_dir),
            "--external",
            str(external_path),
            "--output",
            str(tmp_path / "cli_comparison"),
        ]
    )

    assert exit_code == 0
    payload = json.loads(capsys.readouterr().out)
    assert payload["schema_version"] == FORECAST_COMPARISON_SCHEMA_VERSION
    assert payload["outputs"]["manifest"] == "comparison_manifest.json"
    assert (tmp_path / "cli_comparison" / "forecast_comparison.csv").exists()


def test_compare_cli_missing_external_file_writes_failure_diagnostics(tmp_path, capsys) -> None:
    run_dir = _write_small_run(tmp_path)
    output_dir = tmp_path / "missing_external"
    missing_path = tmp_path / "missing_finance_plan.csv"

    exit_code = main(
        [
            "compare",
            "--run",
            str(run_dir),
            "--external",
            str(missing_path),
            "--output",
            str(output_dir),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "Error:" in captured.err
    payload = json.loads((output_dir / "failure_diagnostics.json").read_text(encoding="utf-8"))
    assert payload["command"] == "compare"
    assert payload["error_type"] == "FileNotFoundError"
    assert str(missing_path) in payload["error"]
    assert (output_dir / "failure_diagnostics.md").exists()


def test_compare_cli_scaffold_model_uses_forecast_long_candidate(tmp_path, capsys) -> None:
    run_dir = _write_small_run(tmp_path)
    forecast_long = pd.read_csv(run_dir / "forecast_long.csv")
    scaffold_model = str(forecast_long["model"].iloc[0])
    model_rows = forecast_long[forecast_long["model"].astype(str) == scaffold_model][["unique_id", "ds", "yhat"]].copy()
    external_path = tmp_path / "finance_plan.csv"
    model_rows.assign(yhat=model_rows["yhat"] + 2).to_csv(external_path, index=False)

    exit_code = main(
        [
            "compare",
            "--run",
            str(run_dir),
            "--external",
            str(external_path),
            "--scaffold-model",
            scaffold_model,
            "--output",
            str(tmp_path / "model_comparison"),
        ]
    )

    assert exit_code == 0
    payload = json.loads(capsys.readouterr().out)
    assert payload["inputs"]["scaffold_forecast_file"] == "forecast_long.csv"
    assert payload["inputs"]["scaffold_model"] == scaffold_model
    comparison = pd.read_csv(tmp_path / "model_comparison" / "forecast_comparison.csv")
    assert set(comparison["scaffold_model"]) == {scaffold_model}


def test_package_root_exports_comparison_helpers() -> None:
    import nixtla_scaffold as ns

    assert ns.FORECAST_COMPARISON_SCHEMA_VERSION == FORECAST_COMPARISON_SCHEMA_VERSION
    assert ns.compare_forecasts is compare_forecasts
    assert ns.write_forecast_comparison is write_forecast_comparison


def _write_small_run(tmp_path):
    history = pd.DataFrame(
        {
            "unique_id": ["Revenue"] * 8,
            "ds": pd.date_range("2025-01-31", periods=8, freq="ME"),
            "y": [100, 105, 108, 112, 118, 121, 127, 130],
        }
    )
    run = run_forecast(history, ForecastSpec(horizon=2, model_policy="baseline"))
    return run.to_directory(tmp_path / "run")
